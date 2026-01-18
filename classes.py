import random
import SimpleGraphics


from math import *
import openpyxl
from scipy.spatial import KDTree


tableau_list_points = "ListPoint.xlsx"
wb = openpyxl.Workbook()
File_points = openpyxl.load_workbook(tableau_list_points)
Sheet_points = File_points.active


class Point:
    def __init__(self, position, color):
        self.position = position  # (x, y)
        self.color = color
       

class PointCloud:
    def __init__(self):
        # Dictionary: Point object -> its coordinates
        self.List_points = {}

    def GetPoints(self, sheet_tableau):
        for row in sheet_tableau.iter_rows(min_row=2, values_only=True):
            if row[1] is not None and row[2] is not None:
                self.addPoint((row[1], row[2]), "red")
            

    def addPoint(self, position, color, ):
        p = Point(position, color)
        self.List_points[p] = position
        x, y = position
        self.drawSinglePoint(x, y, color)
        
        Sheet_points.append([len(self.List_points), x, y])
        Sheet_points.parent.save(tableau_list_points)

    def drawAllPoints(self):
        for p in self.List_points:
            x, y = p.position
            self.drawSinglePoint(x, y, p.color)

    @staticmethod
    def drawSinglePoint(x, y, color):
        SimpleGraphics.setFill(color)
        SimpleGraphics.circle(x, y, 10)

    def createRandomPoints(self, number_of_points, bound_x, bound_y):
        SimpleGraphics.setAutoUpdate(False)
        for _ in range(number_of_points):
            pos = (
                random.randrange(0, bound_x),
                random.randrange(0, bound_y)
            )
            self.addPoint(pos, "red")
        SimpleGraphics.setAutoUpdate(True)

    def getPositions(self):
        # Just return all positions as a list
        return list(self.List_points.values())

    def OptimizeCloud(self):
        points = list(self.List_points.keys())  # List of Point objects
        coords = list(self.List_points.values())  # List of coordinates
        tree = KDTree(coords)
        TempDict = {}
        for i, p in enumerate(points):
            # Query the 2 nearest points: itself + nearest neighbor
            distance, idx = tree.query(p.position, k=2)
            nearest_point = points[idx[1]]  # Nearest OTHER point
            nearest_distance = distance[1]  # Distance to nearest OTHER point
            TempDict[(p.position)] = (idx, nearest_point, nearest_distance)
        print(TempDict)
