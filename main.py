# Projet pour Math en Jeans 2026
# 2 Objectif
#
# n°1 : Lire coordonnées prédifinie dans un tableau excele et l'afficher dans une fenêtre graphique 
# puis placer le centre à l'aide des méthodes
# n°2 : Placement aléatoire de points selon les paramètres donnéss
#
#
# utilisation de libary : 
# simplegraphics-python et openpyxl
#
from time import sleep

from moyenneDesPoints import *
import openpyxl


from classes import *
from Baricentre import *
from PointDejaPresent import *

# initialisation #
screen_width= 1200
screen_height = 900

tableau_excel = "EntreeDesPoints.xlsx"
tableau_list_points = "ListPoint.xlsx"


wb = openpyxl.Workbook()
File_tableau = openpyxl.load_workbook(tableau_excel)
File_points = openpyxl.load_workbook(tableau_list_points)

Sheet_tableau = File_tableau.active
Sheet_points = File_points.active


SimpleGraphics.resize(screen_width, screen_height)
nuage = PointCloud()

def Input():

    print("""--------------------------------\n Séléctionner d'abord un paramètre Basique avant d'utiliser une méthode\n \n 
          1 - Basique : Afficher des points aléatoirement \n
          2 - Basique : afficher les points du tableau \n \n 
          3 - Méthode : afficher la moyenne de ces points \n 
          4 - Méthode : afficher le baricentre de ces points \n \n
          10 - Système : Nettoyer les points \n 
          11- la méthode du point deja présent \n 
          12 - Système : Quitter\n -------------------------------""")
    while True:
            reponse = str(input("Séléctionner un paramètre : "))

            if reponse == "1":
                nombre_de_points = int(input("Combien de points voulez-vous afficher ? : "))
                nuage.createRandomPoints(nombre_de_points,screen_width,screen_height)
                setup()

            elif reponse == "2":
                  nuage.GetPoints(Sheet_tableau)
                  setup()

            elif reponse == "3":
                  Moyenne(nuage)
                  setup()

            elif reponse == "4":
                  baricentre(nuage)
                  setup()

            elif reponse == "10":
                  SimpleGraphics.clear()
                  Sheet_points.delete_rows(2, Sheet_points.max_row)
                  Sheet_points.parent.save(tableau_list_points)


            elif reponse == "11":
                PointDejaPresent(nuage)
                setup()

            elif reponse == "12":
                  print("Fermeture du programme...")
                  SimpleGraphics.close()
                  exit()
def setup():
        nuage.drawAllPoints()
        nuage.getPositions()

if __name__ == "__main__":
    random.seed()
    Input()
